from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime, date, timedelta, timezone
import asyncpg

from database import init_db, close_db, get_pool
from models import (
    Empresa, EmpresaCreate, EmpresaUpdate,
    Moneda, MonedaCreate,
    Categoria, CategoriaCreate, CategoriaUpdate,
    CentroCosto, CentroCostoCreate,
    LineaNegocio, LineaNegocioCreate,
    CuentaFinanciera, CuentaFinancieraCreate, CuentaFinancieraUpdate,
    Tercero, TerceroCreate, TerceroUpdate,
    EmpleadoDetalle, EmpleadoDetalleCreate,
    ArticuloRef, ArticuloRefCreate,
    OC, OCCreate, OCUpdate, OCLinea,
    FacturaProveedor, FacturaProveedorCreate, FacturaProveedorUpdate, FacturaLinea,
    Pago, PagoCreate, PagoDetalle, PagoAplicacion,
    Letra, LetraCreate, LetraUpdate, GenerarLetrasRequest,
    Gasto, GastoCreate, GastoLinea,
    Planilla, PlanillaCreate, PlanillaDetalle,
    Adelanto, AdelantoCreate,
    VentaPOS, CXC, CXCCreate,
    Presupuesto, PresupuestoCreate,
    Conciliacion, ConciliacionCreate, BancoMovRaw, BancoMov,
    DashboardKPIs
)
from odoo_service import OdooService

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Finanzas 4.0 API", version="1.0.0")

api_router = APIRouter(prefix="/api")

# =====================
# STARTUP / SHUTDOWN
# =====================
@app.on_event("startup")
async def startup():
    logger.info("Starting Finanzas 4.0 API...")
    await init_db()
    await seed_data()
    logger.info("Finanzas 4.0 API started successfully")

@app.on_event("shutdown")
async def shutdown():
    await close_db()
    logger.info("Finanzas 4.0 API shutdown complete")

async def seed_data():
    """Create initial seed data"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if seed data exists
        empresa_count = await conn.fetchval("SELECT COUNT(*) FROM finanzas2.cont_empresa")
        if empresa_count > 0:
            logger.info("Seed data already exists, skipping...")
            return
        
        logger.info("Creating seed data...")
        
        # Insert empresa
        await conn.execute("""
            INSERT INTO finanzas2.cont_empresa (nombre, ruc, direccion, telefono, email)
            VALUES ('Mi Empresa S.A.C.', '20123456789', 'Av. Principal 123, Lima', '01-1234567', 'contacto@miempresa.com')
        """)
        
        # Insert moneda PEN
        await conn.execute("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal)
            VALUES ('PEN', 'Sol Peruano', 'S/', TRUE)
        """)
        
        # Insert moneda USD
        await conn.execute("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal)
            VALUES ('USD', 'Dólar Americano', '$', FALSE)
        """)
        
        # Insert categorías
        await conn.execute("""
            INSERT INTO finanzas2.cont_categoria (codigo, nombre, tipo) VALUES
            ('ING-001', 'Ventas', 'ingreso'),
            ('ING-002', 'Otros Ingresos', 'ingreso'),
            ('EGR-001', 'Compras Mercadería', 'egreso'),
            ('EGR-002', 'Servicios', 'egreso'),
            ('EGR-003', 'Planilla', 'egreso'),
            ('EGR-004', 'Alquileres', 'egreso'),
            ('EGR-005', 'Servicios Públicos', 'egreso'),
            ('EGR-006', 'Otros Gastos', 'egreso')
        """)
        
        # Insert centro de costo
        await conn.execute("""
            INSERT INTO finanzas2.cont_centro_costo (codigo, nombre) VALUES
            ('CC-001', 'Administración'),
            ('CC-002', 'Ventas'),
            ('CC-003', 'Operaciones')
        """)
        
        # Insert línea de negocio
        await conn.execute("""
            INSERT INTO finanzas2.cont_linea_negocio (codigo, nombre) VALUES
            ('LN-001', 'Línea Principal'),
            ('LN-002', 'Línea Secundaria')
        """)
        
        # Get PEN moneda_id
        pen_id = await conn.fetchval("SELECT id FROM finanzas2.cont_moneda WHERE codigo = 'PEN'")
        
        # Insert cuenta bancaria
        await conn.execute("""
            INSERT INTO finanzas2.cont_cuenta_financiera (nombre, tipo, banco, numero_cuenta, moneda_id, saldo_actual)
            VALUES ('Cuenta BCP Soles', 'banco', 'BCP', '191-12345678-0-12', $1, 10000.00)
        """, pen_id)
        
        # Insert caja
        await conn.execute("""
            INSERT INTO finanzas2.cont_cuenta_financiera (nombre, tipo, moneda_id, saldo_actual)
            VALUES ('Caja Chica', 'caja', $1, 500.00)
        """, pen_id)
        
        # Insert proveedor
        await conn.execute("""
            INSERT INTO finanzas2.cont_tercero (tipo_documento, numero_documento, nombre, es_proveedor, terminos_pago_dias)
            VALUES ('RUC', '20987654321', 'Proveedor Demo S.A.C.', TRUE, 30)
        """)
        
        logger.info("Seed data created successfully")

# =====================
# HEALTH CHECK
# =====================
@api_router.get("/")
async def root():
    return {"message": "Finanzas 4.0 API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.fetchval("SELECT 1")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

# =====================
# DASHBOARD
# =====================
@api_router.get("/dashboard/kpis", response_model=DashboardKPIs)
async def get_dashboard_kpis():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Total CxP pendiente
        total_cxp = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_cxp WHERE estado NOT IN ('pagado', 'anulada')
        """) or 0
        
        # Total CxC pendiente
        total_cxc = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_cxc WHERE estado NOT IN ('pagado', 'anulada')
        """) or 0
        
        # Total letras pendientes
        total_letras = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_letra WHERE estado IN ('pendiente', 'parcial')
        """) or 0
        
        # Saldo bancos
        saldo_bancos = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_actual), 0) 
            FROM finanzas2.cont_cuenta_financiera WHERE activo = TRUE
        """) or 0
        
        # Ventas del mes
        inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ventas_mes = await conn.fetchval("""
            SELECT COALESCE(SUM(amount_total), 0) 
            FROM finanzas2.cont_venta_pos 
            WHERE date_order >= $1 AND estado_local != 'descartado'
        """, inicio_mes) or 0
        
        # Gastos del mes
        gastos_mes = await conn.fetchval("""
            SELECT COALESCE(SUM(total), 0) 
            FROM finanzas2.cont_gasto WHERE fecha >= $1
        """, inicio_mes.date()) or 0
        
        # Facturas pendientes
        facturas_pendientes = await conn.fetchval("""
            SELECT COUNT(*) FROM finanzas2.cont_factura_proveedor 
            WHERE estado IN ('pendiente', 'parcial')
        """) or 0
        
        # Letras por vencer (próximos 7 días)
        fecha_limite = datetime.now().date() + timedelta(days=7)
        letras_por_vencer = await conn.fetchval("""
            SELECT COUNT(*) FROM finanzas2.cont_letra 
            WHERE estado IN ('pendiente', 'parcial') AND fecha_vencimiento <= $1
        """, fecha_limite) or 0
        
        return DashboardKPIs(
            total_cxp=float(total_cxp),
            total_cxc=float(total_cxc),
            total_letras_pendientes=float(total_letras),
            saldo_bancos=float(saldo_bancos),
            ventas_mes=float(ventas_mes),
            gastos_mes=float(gastos_mes),
            facturas_pendientes=facturas_pendientes,
            letras_por_vencer=letras_por_vencer
        )

# =====================
# EMPRESAS
# =====================
@api_router.get("/empresas", response_model=List[Empresa])
async def list_empresas():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_empresa ORDER BY nombre")
        return [dict(r) for r in rows]

@api_router.post("/empresas", response_model=Empresa)
async def create_empresa(data: EmpresaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_empresa (nombre, ruc, direccion, telefono, email, logo_url, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            RETURNING *
        """, data.nombre, data.ruc, data.direccion, data.telefono, data.email, data.logo_url, data.activo)
        return dict(row)

@api_router.put("/empresas/{id}", response_model=Empresa)
async def update_empresa(id: int, data: EmpresaUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Build dynamic update
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_empresa SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Empresa not found")
        return dict(row)

@api_router.delete("/empresas/{id}")
async def delete_empresa(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_empresa WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Empresa not found")
        return {"message": "Empresa deleted"}

# =====================
# MONEDAS
# =====================
@api_router.get("/monedas", response_model=List[Moneda])
async def list_monedas():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_moneda ORDER BY codigo")
        return [dict(r) for r in rows]

@api_router.post("/monedas", response_model=Moneda)
async def create_moneda(data: MonedaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal, activo)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING *
        """, data.codigo, data.nombre, data.simbolo, data.es_principal, data.activo)
        return dict(row)

@api_router.delete("/monedas/{id}")
async def delete_moneda(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_moneda WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Moneda not found")
        return {"message": "Moneda deleted"}

# =====================
# CATEGORIAS
# =====================
@api_router.get("/categorias", response_model=List[Categoria])
async def list_categorias(tipo: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        if tipo:
            rows = await conn.fetch(
                "SELECT * FROM finanzas2.cont_categoria WHERE tipo = $1 ORDER BY nombre", tipo
            )
        else:
            rows = await conn.fetch("SELECT * FROM finanzas2.cont_categoria ORDER BY tipo, nombre")
        return [dict(r) for r in rows]

@api_router.post("/categorias", response_model=Categoria)
async def create_categoria(data: CategoriaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_categoria (codigo, nombre, tipo, padre_id, descripcion, activo)
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING *
        """, data.codigo, data.nombre, data.tipo, data.padre_id, data.descripcion, data.activo)
        return dict(row)

@api_router.put("/categorias/{id}", response_model=Categoria)
async def update_categoria(id: int, data: CategoriaUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(id)
        query = f"UPDATE finanzas2.cont_categoria SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Categoria not found")
        return dict(row)

@api_router.delete("/categorias/{id}")
async def delete_categoria(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_categoria WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Categoria not found")
        return {"message": "Categoria deleted"}

# =====================
# CENTROS DE COSTO
# =====================
@api_router.get("/centros-costo", response_model=List[CentroCosto])
async def list_centros_costo():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_centro_costo ORDER BY nombre")
        return [dict(r) for r in rows]

@api_router.post("/centros-costo", response_model=CentroCosto)
async def create_centro_costo(data: CentroCostoCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_centro_costo (codigo, nombre, descripcion, activo)
            VALUES ($1, $2, $3, $4)
            RETURNING *
        """, data.codigo, data.nombre, data.descripcion, data.activo)
        return dict(row)

@api_router.delete("/centros-costo/{id}")
async def delete_centro_costo(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_centro_costo WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Centro de costo not found")
        return {"message": "Centro de costo deleted"}

# =====================
# LINEAS DE NEGOCIO
# =====================
@api_router.get("/lineas-negocio", response_model=List[LineaNegocio])
async def list_lineas_negocio():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_linea_negocio ORDER BY nombre")
        return [dict(r) for r in rows]

@api_router.post("/lineas-negocio", response_model=LineaNegocio)
async def create_linea_negocio(data: LineaNegocioCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_linea_negocio (codigo, nombre, descripcion, activo)
            VALUES ($1, $2, $3, $4)
            RETURNING *
        """, data.codigo, data.nombre, data.descripcion, data.activo)
        return dict(row)

@api_router.delete("/lineas-negocio/{id}")
async def delete_linea_negocio(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_linea_negocio WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Línea de negocio not found")
        return {"message": "Línea de negocio deleted"}

# =====================
# CUENTAS FINANCIERAS
# =====================
@api_router.get("/cuentas-financieras", response_model=List[CuentaFinanciera])
async def list_cuentas_financieras(tipo: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        query = """
            SELECT cf.*, m.codigo as moneda_codigo
            FROM finanzas2.cont_cuenta_financiera cf
            LEFT JOIN finanzas2.cont_moneda m ON cf.moneda_id = m.id
        """
        if tipo:
            query += " WHERE cf.tipo = $1"
            rows = await conn.fetch(query + " ORDER BY cf.nombre", tipo)
        else:
            rows = await conn.fetch(query + " ORDER BY cf.nombre")
        return [dict(r) for r in rows]

@api_router.post("/cuentas-financieras", response_model=CuentaFinanciera)
async def create_cuenta_financiera(data: CuentaFinancieraCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_cuenta_financiera 
            (nombre, tipo, banco, numero_cuenta, cci, moneda_id, saldo_actual, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            RETURNING *
        """, data.nombre, data.tipo, data.banco, data.numero_cuenta, data.cci, 
            data.moneda_id, data.saldo_actual, data.activo)
        return dict(row)

@api_router.put("/cuentas-financieras/{id}", response_model=CuentaFinanciera)
async def update_cuenta_financiera(id: int, data: CuentaFinancieraUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(id)
        query = f"UPDATE finanzas2.cont_cuenta_financiera SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Cuenta financiera not found")
        return dict(row)

@api_router.delete("/cuentas-financieras/{id}")
async def delete_cuenta_financiera(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_cuenta_financiera WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Cuenta financiera not found")
        return {"message": "Cuenta financiera deleted"}

# =====================
# TERCEROS (Proveedores, Clientes, Personal)
# =====================
@api_router.get("/terceros", response_model=List[Tercero])
async def list_terceros(
    es_cliente: Optional[bool] = None,
    es_proveedor: Optional[bool] = None,
    es_personal: Optional[bool] = None,
    search: Optional[str] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["activo = TRUE"]
        params = []
        idx = 1
        
        if es_cliente is not None:
            conditions.append(f"es_cliente = ${idx}")
            params.append(es_cliente)
            idx += 1
        if es_proveedor is not None:
            conditions.append(f"es_proveedor = ${idx}")
            params.append(es_proveedor)
            idx += 1
        if es_personal is not None:
            conditions.append(f"es_personal = ${idx}")
            params.append(es_personal)
            idx += 1
        if search:
            conditions.append(f"(nombre ILIKE ${idx} OR numero_documento ILIKE ${idx})")
            params.append(f"%{search}%")
            idx += 1
        
        query = f"SELECT * FROM finanzas2.cont_tercero WHERE {' AND '.join(conditions)} ORDER BY nombre"
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.get("/terceros/{id}", response_model=Tercero)
async def get_tercero(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_tercero WHERE id = $1", id)
        if not row:
            raise HTTPException(404, "Tercero not found")
        return dict(row)

@api_router.post("/terceros", response_model=Tercero)
async def create_tercero(data: TerceroCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_tercero 
            (tipo_documento, numero_documento, nombre, nombre_comercial, direccion, telefono, email,
             es_cliente, es_proveedor, es_personal, terminos_pago_dias, limite_credito, notas, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14)
            RETURNING *
        """, data.tipo_documento, data.numero_documento, data.nombre, data.nombre_comercial,
            data.direccion, data.telefono, data.email, data.es_cliente, data.es_proveedor,
            data.es_personal, data.terminos_pago_dias, data.limite_credito, data.notas, data.activo)
        return dict(row)

@api_router.put("/terceros/{id}", response_model=Tercero)
async def update_tercero(id: int, data: TerceroUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(id)
        query = f"UPDATE finanzas2.cont_tercero SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Tercero not found")
        return dict(row)

@api_router.delete("/terceros/{id}")
async def delete_tercero(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("UPDATE finanzas2.cont_tercero SET activo = FALSE WHERE id = $1", id)
        if result == "UPDATE 0":
            raise HTTPException(404, "Tercero not found")
        return {"message": "Tercero deactivated"}

# Aliases for separated views
@api_router.get("/proveedores", response_model=List[Tercero])
async def list_proveedores(search: Optional[str] = None):
    return await list_terceros(es_proveedor=True, search=search)

@api_router.get("/clientes", response_model=List[Tercero])
async def list_clientes(search: Optional[str] = None):
    return await list_terceros(es_cliente=True, search=search)

@api_router.get("/empleados", response_model=List[Tercero])
async def list_empleados(search: Optional[str] = None):
    return await list_terceros(es_personal=True, search=search)

# =====================
# ARTICULOS REF
# =====================
@api_router.get("/articulos", response_model=List[ArticuloRef])
async def list_articulos(search: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # First try to get from public.prod_inventario if exists
        try:
            inv_rows = await conn.fetch("""
                SELECT id, id as prod_inventario_id, 
                       COALESCE(codigo, '') as codigo,
                       COALESCE(nombre, descripcion, 'Sin nombre') as nombre,
                       descripcion,
                       COALESCE(precio_venta, 0) as precio_referencia,
                       TRUE as activo,
                       NOW() as created_at
                FROM public.prod_inventario
                WHERE ($1::text IS NULL OR nombre ILIKE $1 OR codigo ILIKE $1)
                LIMIT 100
            """, f"%{search}%" if search else None)
            if inv_rows:
                return [dict(r) for r in inv_rows]
        except Exception as e:
            logger.warning(f"Could not fetch from prod_inventario: {e}")
        
        # Fallback to local articulo_ref
        query = "SELECT * FROM finanzas2.cont_articulo_ref WHERE activo = TRUE"
        if search:
            query += " AND (nombre ILIKE $1 OR codigo ILIKE $1)"
            rows = await conn.fetch(query + " ORDER BY nombre LIMIT 100", f"%{search}%")
        else:
            rows = await conn.fetch(query + " ORDER BY nombre LIMIT 100")
        return [dict(r) for r in rows]

# =====================
# INVENTARIO (public.prod_inventario)
# =====================
@api_router.get("/inventario")
async def list_inventario(search: Optional[str] = None):
    """Get items from public.prod_inventario"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT id, codigo, nombre, descripcion, unidad_medida,
                       COALESCE(precio_ref, 0) as precio_ref,
                       COALESCE(costo_compra, 0) as costo_compra,
                       modelo, marca
                FROM public.prod_inventario
                WHERE activo = TRUE
                AND ($1::text IS NULL OR nombre ILIKE $1 OR codigo ILIKE $1 OR descripcion ILIKE $1)
                ORDER BY nombre
                LIMIT 200
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching inventario: {e}")
            return []

# =====================
# MODELOS/CORTES (public.prod_registros + prod_modelos)
# =====================
@api_router.get("/modelos-cortes")
async def list_modelos_cortes(search: Optional[str] = None):
    """Get modelos/cortes from public.prod_registros joining with prod_modelos"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT r.id, r.n_corte, r.modelo_id, r.estado,
                       m.nombre as modelo_nombre,
                       CONCAT(m.nombre, ' - Corte ', r.n_corte) as display_name
                FROM public.prod_registros r
                LEFT JOIN public.prod_modelos m ON r.modelo_id = m.id
                WHERE ($1::text IS NULL OR m.nombre ILIKE $1 OR r.n_corte ILIKE $1)
                ORDER BY r.fecha_creacion DESC
                LIMIT 200
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching modelos/cortes: {e}")
            return []

@api_router.get("/modelos")
async def list_modelos(search: Optional[str] = None):
    """Get modelos from public.prod_modelos"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT id, nombre
                FROM public.prod_modelos
                WHERE ($1::text IS NULL OR nombre ILIKE $1)
                ORDER BY nombre
                LIMIT 100
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching modelos: {e}")
            return []

@api_router.post("/articulos", response_model=ArticuloRef)
async def create_articulo(data: ArticuloRefCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_articulo_ref 
            (prod_inventario_id, codigo, nombre, descripcion, precio_referencia, activo)
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING *
        """, data.prod_inventario_id, data.codigo, data.nombre, data.descripcion, 
            data.precio_referencia, data.activo)
        return dict(row)

# =====================
# ORDENES DE COMPRA
# =====================
async def generate_oc_number(conn) -> str:
    """Generate next OC number"""
    year = datetime.now().year
    prefix = f"OC-{year}-"
    last = await conn.fetchval(f"""
        SELECT numero FROM finanzas2.cont_oc 
        WHERE numero LIKE '{prefix}%' 
        ORDER BY id DESC LIMIT 1
    """)
    if last:
        num = int(last.split('-')[-1]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"

@api_router.get("/ordenes-compra", response_model=List[OC])
async def list_ordenes_compra(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"oc.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"oc.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if fecha_desde:
            conditions.append(f"oc.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"oc.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_oc oc
            LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY oc.fecha DESC, oc.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            oc_dict = dict(row)
            # Get lines
            lineas = await conn.fetch("""
                SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
            """, row['id'])
            oc_dict['lineas'] = [dict(l) for l in lineas]
            result.append(oc_dict)
        
        return result

@api_router.get("/ordenes-compra/{id}", response_model=OC)
async def get_orden_compra(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_oc oc
            LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
            WHERE oc.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Orden de compra not found")
        
        oc_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
        """, id)
        oc_dict['lineas'] = [dict(l) for l in lineas]
        
        return oc_dict

@api_router.post("/ordenes-compra", response_model=OC)
async def create_orden_compra(data: OCCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = await generate_oc_number(conn)
            
            # Calculate totals
            subtotal = 0
            igv = 0
            for linea in data.lineas:
                linea_subtotal = linea.cantidad * linea.precio_unitario
                subtotal += linea_subtotal
                if linea.igv_aplica:
                    igv += linea_subtotal * 0.18
            total = subtotal + igv
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_oc 
                (numero, fecha, proveedor_id, moneda_id, estado, subtotal, igv, total, notas)
                VALUES ($1, $2, $3, $4, 'borrador', $5, $6, $7, $8)
                RETURNING *
            """, numero, data.fecha, data.proveedor_id, data.moneda_id, subtotal, igv, total, data.notas)
            
            oc_id = row['id']
            
            # Insert lines
            for linea in data.lineas:
                linea_subtotal = linea.cantidad * linea.precio_unitario
                await conn.execute("""
                    INSERT INTO finanzas2.cont_oc_linea 
                    (oc_id, articulo_id, descripcion, cantidad, precio_unitario, igv_aplica, subtotal)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                """, oc_id, linea.articulo_id, linea.descripcion, linea.cantidad, 
                    linea.precio_unitario, linea.igv_aplica, linea_subtotal)
            
            # Get full OC with joins within transaction
            oc_row = await conn.fetchrow("""
                SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_oc oc
                LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
                WHERE oc.id = $1
            """, oc_id)
            
            oc_dict = dict(oc_row)
            lineas_rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
            """, oc_id)
            oc_dict['lineas'] = [dict(l) for l in lineas_rows]
            
            return oc_dict

@api_router.put("/ordenes-compra/{id}", response_model=OC)
async def update_orden_compra(id: int, data: OCUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_oc SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING id"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Orden de compra not found")
        
        return await get_orden_compra(id)

@api_router.delete("/ordenes-compra/{id}")
async def delete_orden_compra(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if OC has been converted to factura
        oc = await conn.fetchrow("SELECT * FROM finanzas2.cont_oc WHERE id = $1", id)
        if not oc:
            raise HTTPException(404, "Orden de compra not found")
        if oc['factura_generada_id']:
            raise HTTPException(400, "Cannot delete OC that has generated a factura")
        
        await conn.execute("DELETE FROM finanzas2.cont_oc WHERE id = $1", id)
        return {"message": "Orden de compra deleted"}

@api_router.post("/ordenes-compra/{id}/generar-factura", response_model=FacturaProveedor)
async def generar_factura_desde_oc(id: int):
    """Generate a factura proveedor from an OC"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get OC
            oc = await conn.fetchrow("""
                SELECT * FROM finanzas2.cont_oc WHERE id = $1
            """, id)
            if not oc:
                raise HTTPException(404, "Orden de compra not found")
            if oc['factura_generada_id']:
                raise HTTPException(400, "Esta OC ya generó una factura")
            
            # Generate factura number
            year = datetime.now().year
            prefix = f"FP-{year}-"
            last = await conn.fetchval(f"""
                SELECT numero FROM finanzas2.cont_factura_proveedor 
                WHERE numero LIKE '{prefix}%' 
                ORDER BY id DESC LIMIT 1
            """)
            if last:
                num = int(last.split('-')[-1]) + 1
            else:
                num = 1
            numero = f"{prefix}{num:05d}"
            
            # Create factura
            factura = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_factura_proveedor 
                (numero, proveedor_id, moneda_id, fecha_factura, fecha_vencimiento, 
                 terminos_dias, tipo_documento, estado, subtotal, igv, total, saldo_pendiente, 
                 notas, oc_origen_id)
                VALUES ($1, $2, $3, $4, $5, $6, 'factura', 'pendiente', $7, $8, $9, $9, $10, $11)
                RETURNING *
            """, numero, oc['proveedor_id'], oc['moneda_id'], datetime.now().date(),
                datetime.now().date() + timedelta(days=30), 30,
                oc['subtotal'], oc['igv'], oc['total'], oc['notas'], id)
            
            factura_id = factura['id']
            
            # Copy lines
            oc_lineas = await conn.fetch("SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1", id)
            for linea in oc_lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_factura_proveedor_linea 
                    (factura_id, articulo_id, descripcion, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5)
                """, factura_id, linea['articulo_id'], linea['descripcion'], 
                    linea['subtotal'], linea['igv_aplica'])
            
            # Create CxP
            await conn.execute("""
                INSERT INTO finanzas2.cont_cxp 
                (factura_id, proveedor_id, monto_original, saldo_pendiente, fecha_vencimiento, estado)
                VALUES ($1, $2, $3, $3, $4, 'pendiente')
            """, factura_id, oc['proveedor_id'], oc['total'], 
                datetime.now().date() + timedelta(days=30))
            
            # Update OC
            await conn.execute("""
                UPDATE finanzas2.cont_oc SET estado = 'facturada', factura_generada_id = $1 WHERE id = $2
            """, factura_id, id)
            
            # Get full factura within transaction
            f_row = await conn.fetchrow("""
                SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
                FROM finanzas2.cont_factura_proveedor fp
                LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
                WHERE fp.id = $1
            """, factura_id)
            
            factura_dict = dict(f_row)
            f_lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, factura_id)
            factura_dict['lineas'] = [dict(l) for l in f_lineas]
            
            return factura_dict

# =====================
# FACTURAS PROVEEDOR
# =====================
async def generate_factura_number(conn) -> str:
    year = datetime.now().year
    prefix = f"FP-{year}-"
    last = await conn.fetchval(f"""
        SELECT numero FROM finanzas2.cont_factura_proveedor 
        WHERE numero LIKE '{prefix}%' 
        ORDER BY id DESC LIMIT 1
    """)
    if last:
        num = int(last.split('-')[-1]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"

@api_router.get("/facturas-proveedor", response_model=List[FacturaProveedor])
async def list_facturas_proveedor(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"fp.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"fp.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if fecha_desde:
            conditions.append(f"fp.fecha_factura >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"fp.fecha_factura <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY fp.fecha_factura DESC, fp.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            fp_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre, ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
                LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, row['id'])
            fp_dict['lineas'] = [dict(l) for l in lineas]
            result.append(fp_dict)
        
        return result

async def get_factura_proveedor(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE fp.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Factura not found")
        
        fp_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT fpl.*, c.nombre as categoria_nombre, ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
            FROM finanzas2.cont_factura_proveedor_linea fpl
            LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
            LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
            WHERE fpl.factura_id = $1 ORDER BY fpl.id
        """, id)
        fp_dict['lineas'] = [dict(l) for l in lineas]
        
        return fp_dict

@api_router.get("/facturas-proveedor/{id}", response_model=FacturaProveedor)
async def get_factura_proveedor_endpoint(id: int):
    return await get_factura_proveedor(id)

@api_router.post("/facturas-proveedor", response_model=FacturaProveedor)
async def create_factura_proveedor(data: FacturaProveedorCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = data.numero or await generate_factura_number(conn)
            
            # Calculate totals
            subtotal = sum(l.importe for l in data.lineas)
            if data.impuestos_incluidos:
                subtotal = subtotal / 1.18
                igv = subtotal * 0.18
            else:
                igv = sum(l.importe * 0.18 for l in data.lineas if l.igv_aplica)
            total = subtotal + igv
            
            # Calculate fecha_vencimiento
            fecha_vencimiento = data.fecha_vencimiento
            if not fecha_vencimiento and data.terminos_dias:
                fecha_vencimiento = data.fecha_factura + timedelta(days=data.terminos_dias)
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_factura_proveedor 
                (numero, proveedor_id, beneficiario_nombre, moneda_id, fecha_factura, fecha_vencimiento,
                 terminos_dias, tipo_documento, estado, subtotal, igv, total, saldo_pendiente, 
                 impuestos_incluidos, notas)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, 'pendiente', $9, $10, $11, $11, $12, $13)
                RETURNING id
            """, numero, data.proveedor_id, data.beneficiario_nombre, data.moneda_id,
                data.fecha_factura, fecha_vencimiento, data.terminos_dias, data.tipo_documento,
                subtotal, igv, total, data.impuestos_incluidos, data.notas)
            
            factura_id = row['id']
            
            # Insert lines
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_factura_proveedor_linea 
                    (factura_id, categoria_id, articulo_id, descripcion, linea_negocio_id, 
                     centro_costo_id, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                """, factura_id, linea.categoria_id, linea.articulo_id, linea.descripcion,
                    linea.linea_negocio_id, linea.centro_costo_id, linea.importe, linea.igv_aplica)
            
            # Create CxP
            await conn.execute("""
                INSERT INTO finanzas2.cont_cxp 
                (factura_id, proveedor_id, monto_original, saldo_pendiente, fecha_vencimiento, estado)
                VALUES ($1, $2, $3, $3, $4, 'pendiente')
            """, factura_id, data.proveedor_id, total, fecha_vencimiento)
            
            # Get the created factura with relations within the same transaction
            row = await conn.fetchrow("""
                SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
                FROM finanzas2.cont_factura_proveedor fp
                LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
                WHERE fp.id = $1
            """, factura_id)
            
            if not row:
                raise HTTPException(404, "Factura not found after creation")
            
            fp_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre, ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
                LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, factura_id)
            fp_dict['lineas'] = [dict(l) for l in lineas]
            
            return fp_dict

@api_router.put("/facturas-proveedor/{id}", response_model=FacturaProveedor)
async def update_factura_proveedor(id: int, data: FacturaProveedorUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if factura can be edited
        factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1", id)
        if not factura:
            raise HTTPException(404, "Factura not found")
        if factura['estado'] in ('pagado', 'anulada'):
            raise HTTPException(400, "Cannot edit paid or cancelled factura")
        
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_factura_proveedor SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx}"
        await conn.execute(query, *values)
        
        return await get_factura_proveedor(id)

@api_router.delete("/facturas-proveedor/{id}")
async def delete_factura_proveedor(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1", id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            
            # Check if has payments or letras
            pagos = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion 
                WHERE tipo_documento = 'factura' AND documento_id = $1
            """, id)
            if pagos > 0:
                raise HTTPException(400, "Cannot delete factura with payments. Reverse payments first.")
            
            letras = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, id)
            if letras > 0:
                raise HTTPException(400, "Cannot delete factura with letras. Delete letras first.")
            
            # Delete CxP and factura
            await conn.execute("DELETE FROM finanzas2.cont_cxp WHERE factura_id = $1", id)
            await conn.execute("DELETE FROM finanzas2.cont_factura_proveedor WHERE id = $1", id)
            
            return {"message": "Factura deleted"}

@api_router.get("/facturas-proveedor/{id}/pagos")
async def get_pagos_de_factura(id: int):
    """Get all payments applied to a specific factura"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Get pagos through pago_aplicacion
        rows = await conn.fetch("""
            SELECT p.*, pa.monto_aplicado, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_pago p
            JOIN finanzas2.cont_pago_aplicacion pa ON pa.pago_id = p.id
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            WHERE pa.tipo_documento = 'factura' AND pa.documento_id = $1
            ORDER BY p.fecha DESC
        """, id)
        
        return [dict(r) for r in rows]

@api_router.get("/facturas-proveedor/{id}/letras")
async def get_letras_de_factura(id: int):
    """Get all letras generated from a specific factura"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT l.*, 
                   fp.moneda_id,
                   m.codigo as moneda_codigo, 
                   m.simbolo as moneda_simbolo
            FROM finanzas2.cont_letra l
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON l.factura_id = fp.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE l.factura_id = $1
            ORDER BY l.fecha_vencimiento ASC
        """, id)
        
        return [dict(r) for r in rows]

@api_router.post("/facturas-proveedor/{id}/deshacer-canje")
async def deshacer_canje_letras(id: int):
    """Reverse the canje of letras - delete all letras and set factura back to pendiente"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1", id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            
            if factura['estado'] != 'canjeado':
                raise HTTPException(400, "Factura is not in canjeado state")
            
            # Check if any letra has payments
            pagos_letras = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion pa
                JOIN finanzas2.cont_letra l ON pa.tipo_documento = 'letra' AND pa.documento_id = l.id
                WHERE l.factura_id = $1
            """, id)
            
            if pagos_letras > 0:
                raise HTTPException(400, "Cannot undo canje - some letras have payments. Delete payments first.")
            
            # Delete all letras
            await conn.execute("DELETE FROM finanzas2.cont_letra WHERE factura_id = $1", id)
            
            # Update factura back to pendiente
            await conn.execute("""
                UPDATE finanzas2.cont_factura_proveedor 
                SET estado = 'pendiente', updated_at = NOW() 
                WHERE id = $1
            """, id)
            
            return {"message": "Canje reversed successfully"}

# =====================
# PAGOS
# =====================
async def generate_pago_number(conn, tipo: str) -> str:
    year = datetime.now().year
    prefix = f"PAG-{tipo[0].upper()}-{year}-"
    last = await conn.fetchval(f"""
        SELECT numero FROM finanzas2.cont_pago 
        WHERE numero LIKE '{prefix}%' 
        ORDER BY id DESC LIMIT 1
    """)
    if last:
        num = int(last.split('-')[-1]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"

@api_router.get("/pagos", response_model=List[Pago])
async def list_pagos(
    tipo: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if tipo:
            conditions.append(f"p.tipo = ${idx}")
            params.append(tipo)
            idx += 1
        if fecha_desde:
            conditions.append(f"p.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"p.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY p.fecha DESC, p.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            pago_dict = dict(row)
            
            # Get detalles
            detalles = await conn.fetch("""
                SELECT pd.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_pago_detalle pd
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
                WHERE pd.pago_id = $1
            """, row['id'])
            pago_dict['detalles'] = [dict(d) for d in detalles]
            
            # Get aplicaciones
            aplicaciones = await conn.fetch("""
                SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
            """, row['id'])
            pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
            
            result.append(pago_dict)
        
        return result

@api_router.post("/pagos", response_model=Pago)
async def create_pago(data: PagoCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Validate that payment amount doesn't exceed document balance
            for aplicacion in data.aplicaciones:
                if aplicacion.tipo_documento == 'factura':
                    doc = await conn.fetchrow("""
                        SELECT saldo_pendiente, total, estado FROM finanzas2.cont_factura_proveedor WHERE id = $1
                    """, aplicacion.documento_id)
                    if not doc:
                        raise HTTPException(404, f"Factura {aplicacion.documento_id} not found")
                    if doc['estado'] == 'canjeado':
                        raise HTTPException(400, "No se puede pagar una factura canjeada. Debe pagar las letras.")
                    if aplicacion.monto_aplicado > float(doc['saldo_pendiente']):
                        raise HTTPException(400, f"El monto ({aplicacion.monto_aplicado:.2f}) excede el saldo pendiente ({doc['saldo_pendiente']:.2f})")
                elif aplicacion.tipo_documento == 'letra':
                    doc = await conn.fetchrow("""
                        SELECT saldo_pendiente, monto, estado FROM finanzas2.cont_letra WHERE id = $1
                    """, aplicacion.documento_id)
                    if not doc:
                        raise HTTPException(404, f"Letra {aplicacion.documento_id} not found")
                    if aplicacion.monto_aplicado > float(doc['saldo_pendiente']):
                        raise HTTPException(400, f"El monto ({aplicacion.monto_aplicado:.2f}) excede el saldo pendiente ({doc['saldo_pendiente']:.2f})")
            
            numero = await generate_pago_number(conn, data.tipo)
            
            # Create pago
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, referencia, notas)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                RETURNING *
            """, numero, data.tipo, data.fecha, data.cuenta_financiera_id, data.moneda_id,
                data.monto_total, data.referencia, data.notas)
            
            pago_id = pago['id']
            
            # Insert detalles (multi-medio)
            for detalle in data.detalles:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (pago_id, cuenta_financiera_id, medio_pago, monto, referencia)
                    VALUES ($1, $2, $3, $4, $5)
                """, pago_id, detalle.cuenta_financiera_id, detalle.medio_pago, 
                    detalle.monto, detalle.referencia)
                
                # Update cuenta financiera saldo
                if data.tipo == 'egreso':
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual - $1 WHERE id = $2
                    """, detalle.monto, detalle.cuenta_financiera_id)
                else:
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual + $1 WHERE id = $2
                    """, detalle.monto, detalle.cuenta_financiera_id)
            
            # Insert aplicaciones and update documents
            for aplicacion in data.aplicaciones:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_aplicacion 
                    (pago_id, tipo_documento, documento_id, monto_aplicado)
                    VALUES ($1, $2, $3, $4)
                """, pago_id, aplicacion.tipo_documento, aplicacion.documento_id, 
                    aplicacion.monto_aplicado)
                
                # Update document saldo
                if aplicacion.tipo_documento == 'factura':
                    await conn.execute("""
                        UPDATE finanzas2.cont_factura_proveedor 
                        SET saldo_pendiente = saldo_pendiente - $1 WHERE id = $2
                    """, aplicacion.monto_aplicado, aplicacion.documento_id)
                    
                    # Update estado
                    fp = await conn.fetchrow("""
                        SELECT total, saldo_pendiente FROM finanzas2.cont_factura_proveedor WHERE id = $1
                    """, aplicacion.documento_id)
                    if fp['saldo_pendiente'] <= 0:
                        await conn.execute("""
                            UPDATE finanzas2.cont_factura_proveedor SET estado = 'pagado' WHERE id = $1
                        """, aplicacion.documento_id)
                        await conn.execute("""
                            UPDATE finanzas2.cont_cxp SET estado = 'pagado', saldo_pendiente = 0 WHERE factura_id = $1
                        """, aplicacion.documento_id)
                    else:
                        await conn.execute("""
                            UPDATE finanzas2.cont_factura_proveedor SET estado = 'parcial' WHERE id = $1
                        """, aplicacion.documento_id)
                        await conn.execute("""
                            UPDATE finanzas2.cont_cxp SET estado = 'parcial', saldo_pendiente = $2 WHERE factura_id = $1
                        """, aplicacion.documento_id, fp['saldo_pendiente'])
                
                elif aplicacion.tipo_documento == 'letra':
                    await conn.execute("""
                        UPDATE finanzas2.cont_letra 
                        SET saldo_pendiente = saldo_pendiente - $1 WHERE id = $2
                    """, aplicacion.monto_aplicado, aplicacion.documento_id)
                    
                    letra = await conn.fetchrow("""
                        SELECT monto, saldo_pendiente FROM finanzas2.cont_letra WHERE id = $1
                    """, aplicacion.documento_id)
                    if letra['saldo_pendiente'] <= 0:
                        await conn.execute("""
                            UPDATE finanzas2.cont_letra SET estado = 'pagada' WHERE id = $1
                        """, aplicacion.documento_id)
                    else:
                        await conn.execute("""
                            UPDATE finanzas2.cont_letra SET estado = 'parcial' WHERE id = $1
                        """, aplicacion.documento_id)
            
            # Get full pago with relations within the same transaction
            row = await conn.fetchrow("""
                SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_pago p
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
                LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
                WHERE p.id = $1
            """, pago_id)
            
            if not row:
                raise HTTPException(404, "Pago not found after creation")
            
            pago_dict = dict(row)
            
            detalles = await conn.fetch("""
                SELECT pd.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_pago_detalle pd
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
                WHERE pd.pago_id = $1
            """, pago_id)
            pago_dict['detalles'] = [dict(d) for d in detalles]
            
            aplicaciones = await conn.fetch("""
                SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
            """, pago_id)
            pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
            
            return pago_dict

async def get_pago(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            WHERE p.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Pago not found")
        
        pago_dict = dict(row)
        
        detalles = await conn.fetch("""
            SELECT pd.*, cf.nombre as cuenta_nombre
            FROM finanzas2.cont_pago_detalle pd
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
            WHERE pd.pago_id = $1
        """, id)
        pago_dict['detalles'] = [dict(d) for d in detalles]
        
        aplicaciones = await conn.fetch("""
            SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
        """, id)
        pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
        
        return pago_dict

@api_router.get("/pagos/{id}", response_model=Pago)
async def get_pago_endpoint(id: int):
    return await get_pago(id)

@api_router.delete("/pagos/{id}")
async def delete_pago(id: int):
    """Delete a payment and reverse all its effects"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            pago = await conn.fetchrow("SELECT * FROM finanzas2.cont_pago WHERE id = $1", id)
            if not pago:
                raise HTTPException(404, "Pago not found")
            
            # Reverse cuenta financiera updates
            detalles = await conn.fetch("SELECT * FROM finanzas2.cont_pago_detalle WHERE pago_id = $1", id)
            for detalle in detalles:
                if pago['tipo'] == 'egreso':
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual + $1 WHERE id = $2
                    """, detalle['monto'], detalle['cuenta_financiera_id'])
                else:
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual - $1 WHERE id = $2
                    """, detalle['monto'], detalle['cuenta_financiera_id'])
            
            # Reverse aplicaciones
            aplicaciones = await conn.fetch("SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1", id)
            for aplicacion in aplicaciones:
                if aplicacion['tipo_documento'] == 'factura':
                    await conn.execute("""
                        UPDATE finanzas2.cont_factura_proveedor 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
                    await conn.execute("""
                        UPDATE finanzas2.cont_cxp 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE factura_id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
                elif aplicacion['tipo_documento'] == 'letra':
                    await conn.execute("""
                        UPDATE finanzas2.cont_letra 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
            
            # Delete pago
            await conn.execute("DELETE FROM finanzas2.cont_pago WHERE id = $1", id)
            
            return {"message": "Pago deleted and reversed"}

# =====================
# LETRAS
# =====================
@api_router.get("/letras", response_model=List[Letra])
async def list_letras(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    factura_id: Optional[int] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"l.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"l.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if factura_id:
            conditions.append(f"l.factura_id = ${idx}")
            params.append(factura_id)
            idx += 1
        
        query = f"""
            SELECT l.*, t.nombre as proveedor_nombre, fp.numero as factura_numero
            FROM finanzas2.cont_letra l
            LEFT JOIN finanzas2.cont_tercero t ON l.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON l.factura_id = fp.id
            WHERE {' AND '.join(conditions)}
            ORDER BY l.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/letras/generar", response_model=List[Letra])
async def generar_letras(data: GenerarLetrasRequest):
    """Generate letras from a factura proveedor"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get factura
            factura = await conn.fetchrow("""
                SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1
            """, data.factura_id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            if factura['estado'] in ('pagado', 'anulada', 'canjeado'):
                raise HTTPException(400, "Cannot generate letras for this factura")
            
            # Check if already has letras
            existing = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, data.factura_id)
            if existing > 0:
                raise HTTPException(400, "Factura already has letras")
            
            letras = []
            
            # Check if using custom letras
            if data.letras_personalizadas and len(data.letras_personalizadas) > 0:
                # Validate total matches factura total
                total_letras = sum(l.monto for l in data.letras_personalizadas)
                if abs(total_letras - float(factura['total'])) > 0.01:
                    raise HTTPException(400, f"El total de las letras ({total_letras:.2f}) debe ser igual al total de la factura ({factura['total']:.2f})")
                
                for i, letra_data in enumerate(data.letras_personalizadas):
                    numero = f"L-{factura['numero']}-{i+1:02d}"
                    
                    letra = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_letra 
                        (numero, factura_id, proveedor_id, monto, fecha_emision, fecha_vencimiento, 
                         estado, saldo_pendiente)
                        VALUES ($1, $2, $3, $4, $5, $6, 'pendiente', $4)
                        RETURNING *
                    """, numero, data.factura_id, factura['proveedor_id'], letra_data.monto,
                        datetime.now().date(), letra_data.fecha_vencimiento)
                    letras.append(dict(letra))
            else:
                # Use automatic calculation
                monto_por_letra = data.monto_por_letra or (factura['total'] / data.cantidad_letras)
                fecha_base = factura['fecha_vencimiento'] or datetime.now().date()
                
                for i in range(data.cantidad_letras):
                    fecha_vencimiento = fecha_base + timedelta(days=data.dias_entre_letras * i)
                    numero = f"L-{factura['numero']}-{i+1:02d}"
                    
                    letra = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_letra 
                        (numero, factura_id, proveedor_id, monto, fecha_emision, fecha_vencimiento, 
                         estado, saldo_pendiente)
                        VALUES ($1, $2, $3, $4, $5, $6, 'pendiente', $4)
                        RETURNING *
                    """, numero, data.factura_id, factura['proveedor_id'], monto_por_letra,
                        datetime.now().date(), fecha_vencimiento)
                    letras.append(dict(letra))
            
            # Update factura status to 'canjeado'
            await conn.execute("""
                UPDATE finanzas2.cont_factura_proveedor SET estado = 'canjeado' WHERE id = $1
            """, data.factura_id)
            await conn.execute("""
                UPDATE finanzas2.cont_cxp SET estado = 'canjeado' WHERE factura_id = $1
            """, data.factura_id)
            
            return letras

@api_router.delete("/letras/{id}")
async def delete_letra(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            letra = await conn.fetchrow("SELECT * FROM finanzas2.cont_letra WHERE id = $1", id)
            if not letra:
                raise HTTPException(404, "Letra not found")
            
            # Check if has payments
            pagos = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion 
                WHERE tipo_documento = 'letra' AND documento_id = $1
            """, id)
            if pagos > 0:
                raise HTTPException(400, "Cannot delete letra with payments. Reverse payments first.")
            
            factura_id = letra['factura_id']
            
            await conn.execute("DELETE FROM finanzas2.cont_letra WHERE id = $1", id)
            
            # Check if factura has remaining letras
            remaining = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, factura_id)
            
            if remaining == 0:
                # Revert factura to pendiente
                await conn.execute("""
                    UPDATE finanzas2.cont_factura_proveedor SET estado = 'pendiente' WHERE id = $1
                """, factura_id)
                await conn.execute("""
                    UPDATE finanzas2.cont_cxp SET estado = 'pendiente' WHERE factura_id = $1
                """, factura_id)
            
            return {"message": "Letra deleted"}

# =====================
# GASTOS
# =====================
async def generate_gasto_number(conn) -> str:
    year = datetime.now().year
    prefix = f"GAS-{year}-"
    last = await conn.fetchval(f"""
        SELECT numero FROM finanzas2.cont_gasto 
        WHERE numero LIKE '{prefix}%' 
        ORDER BY id DESC LIMIT 1
    """)
    if last:
        num = int(last.split('-')[-1]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"

@api_router.get("/gastos", response_model=List[Gasto])
async def list_gastos(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if fecha_desde:
            conditions.append(f"g.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"g.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY g.fecha DESC, g.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            gasto_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT gl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_gasto_linea gl
                LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
                WHERE gl.gasto_id = $1 ORDER BY gl.id
            """, row['id'])
            gasto_dict['lineas'] = [dict(l) for l in lineas]
            result.append(gasto_dict)
        
        return result

@api_router.post("/gastos", response_model=Gasto)
async def create_gasto(data: GastoCreate):
    """Create a gasto with mandatory payment(s) that must cover total"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = await generate_gasto_number(conn)
            
            # Calculate totals
            subtotal = sum(l.importe for l in data.lineas)
            igv = sum(l.importe * 0.18 for l in data.lineas if l.igv_aplica)
            total = subtotal + igv
            
            # Validate payments sum to total
            if not data.pagos:
                raise HTTPException(400, "El gasto debe tener al menos un pago")
            
            total_pagos = sum(p.monto for p in data.pagos)
            if abs(total_pagos - total) > 0.01:
                raise HTTPException(400, f"El total de pagos ({total_pagos:.2f}) debe ser igual al total del gasto ({total:.2f})")
            
            # Create gasto first
            gasto = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_gasto 
                (numero, fecha, proveedor_id, beneficiario_nombre, moneda_id, subtotal, igv, total,
                 tipo_documento, numero_documento, notas)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)
                RETURNING id
            """, numero, data.fecha, data.proveedor_id, data.beneficiario_nombre, data.moneda_id,
                subtotal, igv, total, data.tipo_documento, data.numero_documento, data.notas)
            
            gasto_id = gasto['id']
            
            # Insert lineas
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_gasto_linea 
                    (gasto_id, categoria_id, descripcion, linea_negocio_id, centro_costo_id, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                """, gasto_id, linea.categoria_id, linea.descripcion, linea.linea_negocio_id,
                    linea.centro_costo_id, linea.importe, linea.igv_aplica)
            
            # Create pago(s)
            pago_numero = await generate_pago_number(conn, 'egreso')
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, notas)
                VALUES ($1, 'egreso', $2, $3, $4, $5, $6)
                RETURNING id
            """, pago_numero, data.fecha, data.pagos[0].cuenta_financiera_id, data.moneda_id,
                total, f"Pago de gasto {numero}")
            
            pago_id = pago['id']
            
            # Insert pago detalles (multiple methods)
            for pago_det in data.pagos:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (pago_id, cuenta_financiera_id, medio_pago, monto, referencia)
                    VALUES ($1, $2, $3, $4, $5)
                """, pago_id, pago_det.cuenta_financiera_id, pago_det.medio_pago, 
                    pago_det.monto, pago_det.referencia)
                
                # Update cuenta financiera
                await conn.execute("""
                    UPDATE finanzas2.cont_cuenta_financiera 
                    SET saldo_actual = saldo_actual - $1 WHERE id = $2
                """, pago_det.monto, pago_det.cuenta_financiera_id)
            
            # Insert pago aplicacion
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado)
                VALUES ($1, 'gasto', $2, $3)
            """, pago_id, gasto_id, total)
            
            # Get full gasto data within transaction
            row = await conn.fetchrow("""
                SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_gasto g
                LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
                WHERE g.id = $1
            """, gasto_id)
            
            gasto_dict = dict(row)
            lineas_rows = await conn.fetch("""
                SELECT gl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_gasto_linea gl
                LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
                WHERE gl.gasto_id = $1 ORDER BY gl.id
            """, gasto_id)
            gasto_dict['lineas'] = [dict(l) for l in lineas_rows]
            
            return gasto_dict

async def get_gasto(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE g.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Gasto not found")
        
        gasto_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT gl.*, c.nombre as categoria_nombre
            FROM finanzas2.cont_gasto_linea gl
            LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
            WHERE gl.gasto_id = $1 ORDER BY gl.id
        """, id)
        gasto_dict['lineas'] = [dict(l) for l in lineas]
        
        return gasto_dict

@api_router.get("/gastos/{id}", response_model=Gasto)
async def get_gasto_endpoint(id: int):
    return await get_gasto(id)

@api_router.delete("/gastos/{id}")
async def delete_gasto(id: int):
    """Delete a gasto and its associated lines and payments"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if gasto exists
        gasto = await conn.fetchrow("SELECT * FROM finanzas2.cont_gasto WHERE id = $1", id)
        if not gasto:
            raise HTTPException(status_code=404, detail="Gasto no encontrado")
        
        # Delete associated lines
        await conn.execute("DELETE FROM finanzas2.cont_gasto_linea WHERE gasto_id = $1", id)
        
        # Delete the gasto (pagos should cascade or be handled separately)
        await conn.execute("DELETE FROM finanzas2.cont_gasto WHERE id = $1", id)
        
        return {"message": "Gasto eliminado exitosamente"}

# =====================
# ADELANTOS
# =====================
@api_router.get("/adelantos", response_model=List[Adelanto])
async def list_adelantos(
    empleado_id: Optional[int] = None,
    pagado: Optional[bool] = None,
    descontado: Optional[bool] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if empleado_id:
            conditions.append(f"a.empleado_id = ${idx}")
            params.append(empleado_id)
            idx += 1
        if pagado is not None:
            conditions.append(f"a.pagado = ${idx}")
            params.append(pagado)
            idx += 1
        if descontado is not None:
            conditions.append(f"a.descontado = ${idx}")
            params.append(descontado)
            idx += 1
        
        query = f"""
            SELECT a.*, t.nombre as empleado_nombre
            FROM finanzas2.cont_adelanto_empleado a
            LEFT JOIN finanzas2.cont_tercero t ON a.empleado_id = t.id
            WHERE {' AND '.join(conditions)}
            ORDER BY a.fecha DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/adelantos", response_model=Adelanto)
async def create_adelanto(data: AdelantoCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            pago_id = None
            
            if data.pagar and data.cuenta_financiera_id:
                # Create pago
                pago_numero = await generate_pago_number(conn, 'egreso')
                pago = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_pago 
                    (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas)
                    VALUES ($1, 'egreso', $2, $3, $4, $5)
                    RETURNING id
                """, pago_numero, data.fecha, data.cuenta_financiera_id, data.monto, 
                    "Adelanto a empleado")
                pago_id = pago['id']
                
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (pago_id, cuenta_financiera_id, medio_pago, monto)
                    VALUES ($1, $2, $3, $4)
                """, pago_id, data.cuenta_financiera_id, data.medio_pago, data.monto)
                
                await conn.execute("""
                    UPDATE finanzas2.cont_cuenta_financiera 
                    SET saldo_actual = saldo_actual - $1 WHERE id = $2
                """, data.monto, data.cuenta_financiera_id)
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_adelanto_empleado 
                (empleado_id, fecha, monto, motivo, pagado, pago_id)
                VALUES ($1, $2, $3, $4, $5, $6)
                RETURNING *
            """, data.empleado_id, data.fecha, data.monto, data.motivo, 
                data.pagar, pago_id)
            
            if pago_id:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_aplicacion 
                    (pago_id, tipo_documento, documento_id, monto_aplicado)
                    VALUES ($1, 'adelanto', $2, $3)
                """, pago_id, row['id'], data.monto)
            
            # Get empleado nombre
            emp = await conn.fetchrow("SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1", data.empleado_id)
            result = dict(row)
            result['empleado_nombre'] = emp['nombre'] if emp else None
            
            return result

@api_router.post("/adelantos/{id}/pagar", response_model=Adelanto)
async def pagar_adelanto(
    id: int, 
    cuenta_financiera_id: int = Query(...),
    medio_pago: str = Query(default="efectivo")
):
    """Register payment for an existing unpaid advance"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get the adelanto
            adelanto = await conn.fetchrow(
                "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1", id
            )
            if not adelanto:
                raise HTTPException(404, "Adelanto no encontrado")
            if adelanto['pagado']:
                raise HTTPException(400, "Este adelanto ya fue pagado")
            
            # Create pago
            pago_numero = await generate_pago_number(conn, 'egreso')
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas)
                VALUES ($1, 'egreso', CURRENT_DATE, $2, $3, $4)
                RETURNING id
            """, pago_numero, cuenta_financiera_id, adelanto['monto'], 
                "Pago de adelanto a empleado")
            pago_id = pago['id']
            
            # Create pago detalle
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle 
                (pago_id, cuenta_financiera_id, medio_pago, monto)
                VALUES ($1, $2, $3, $4)
            """, pago_id, cuenta_financiera_id, medio_pago, adelanto['monto'])
            
            # Update cuenta saldo
            await conn.execute("""
                UPDATE finanzas2.cont_cuenta_financiera 
                SET saldo_actual = saldo_actual - $1 WHERE id = $2
            """, adelanto['monto'], cuenta_financiera_id)
            
            # Update adelanto
            row = await conn.fetchrow("""
                UPDATE finanzas2.cont_adelanto_empleado 
                SET pagado = TRUE, pago_id = $1
                WHERE id = $2
                RETURNING *
            """, pago_id, id)
            
            # Create pago aplicacion
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado)
                VALUES ($1, 'adelanto', $2, $3)
            """, pago_id, id, adelanto['monto'])
            
            # Get empleado nombre
            emp = await conn.fetchrow(
                "SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1", 
                row['empleado_id']
            )
            result = dict(row)
            result['empleado_nombre'] = emp['nombre'] if emp else None
            
            return result

@api_router.put("/adelantos/{id}", response_model=Adelanto)
async def update_adelanto(id: int, data: AdelantoCreate):
    """Update an existing advance"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if adelanto exists and is not already paid/discounted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1", id
        )
        if not existing:
            raise HTTPException(404, "Adelanto no encontrado")
        if existing['pagado'] or existing['descontado']:
            raise HTTPException(400, "No se puede editar un adelanto pagado o descontado")
        
        row = await conn.fetchrow("""
            UPDATE finanzas2.cont_adelanto_empleado 
            SET empleado_id = $1, fecha = $2, monto = $3, motivo = $4
            WHERE id = $5
            RETURNING *
        """, data.empleado_id, data.fecha, data.monto, data.motivo, id)
        
        # Get empleado nombre
        emp = await conn.fetchrow(
            "SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1", 
            row['empleado_id']
        )
        result = dict(row)
        result['empleado_nombre'] = emp['nombre'] if emp else None
        
        return result

@api_router.delete("/adelantos/{id}")
async def delete_adelanto(id: int):
    """Delete an advance (only if not paid or discounted)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if adelanto exists and can be deleted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1", id
        )
        if not existing:
            raise HTTPException(404, "Adelanto no encontrado")
        if existing['pagado']:
            raise HTTPException(400, "No se puede eliminar un adelanto pagado. Primero anule el pago.")
        if existing['descontado']:
            raise HTTPException(400, "No se puede eliminar un adelanto ya descontado en planilla")
        
        await conn.execute(
            "DELETE FROM finanzas2.cont_adelanto_empleado WHERE id = $1", id
        )
        return {"message": "Adelanto eliminado"}

# =====================
# PLANILLA
# =====================
@api_router.get("/planillas", response_model=List[Planilla])
async def list_planillas():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT * FROM finanzas2.cont_planilla ORDER BY periodo DESC
        """)
        
        result = []
        for row in rows:
            planilla_dict = dict(row)
            detalles = await conn.fetch("""
                SELECT pd.*, t.nombre as empleado_nombre
                FROM finanzas2.cont_planilla_detalle pd
                LEFT JOIN finanzas2.cont_tercero t ON pd.empleado_id = t.id
                WHERE pd.planilla_id = $1
            """, row['id'])
            planilla_dict['detalles'] = [dict(d) for d in detalles]
            result.append(planilla_dict)
        
        return result

@api_router.post("/planillas", response_model=Planilla)
async def create_planilla(data: PlanillaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Calculate totals
            total_bruto = sum(d.salario_base + d.bonificaciones for d in data.detalles)
            total_adelantos = sum(d.adelantos for d in data.detalles)
            total_descuentos = sum(d.otros_descuentos for d in data.detalles)
            total_neto = total_bruto - total_adelantos - total_descuentos
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_planilla 
                (periodo, fecha_inicio, fecha_fin, total_bruto, total_adelantos, 
                 total_descuentos, total_neto, estado)
                VALUES ($1, $2, $3, $4, $5, $6, $7, 'borrador')
                RETURNING *
            """, data.periodo, data.fecha_inicio, data.fecha_fin, total_bruto,
                total_adelantos, total_descuentos, total_neto)
            
            planilla_id = row['id']
            planilla_dict = dict(row)
            detalles_list = []
            
            for detalle in data.detalles:
                neto_pagar = detalle.salario_base + detalle.bonificaciones - detalle.adelantos - detalle.otros_descuentos
                detalle_row = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_planilla_detalle 
                    (planilla_id, empleado_id, salario_base, bonificaciones, adelantos, 
                     otros_descuentos, neto_pagar)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                    RETURNING *
                """, planilla_id, detalle.empleado_id, detalle.salario_base, detalle.bonificaciones,
                    detalle.adelantos, detalle.otros_descuentos, neto_pagar)
                
                # Get employee name
                emp = await conn.fetchrow("SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1", detalle.empleado_id)
                detalle_dict = dict(detalle_row)
                detalle_dict['empleado_nombre'] = emp['nombre'] if emp else None
                detalles_list.append(detalle_dict)
            
            planilla_dict['detalles'] = detalles_list
            return planilla_dict

async def get_planilla(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_planilla WHERE id = $1", id)
        if not row:
            raise HTTPException(404, "Planilla not found")
        
        planilla_dict = dict(row)
        detalles = await conn.fetch("""
            SELECT pd.*, t.nombre as empleado_nombre
            FROM finanzas2.cont_planilla_detalle pd
            LEFT JOIN finanzas2.cont_tercero t ON pd.empleado_id = t.id
            WHERE pd.planilla_id = $1
        """, id)
        planilla_dict['detalles'] = [dict(d) for d in detalles]
        
        return planilla_dict

@api_router.get("/planillas/{id}", response_model=Planilla)
async def get_planilla_endpoint(id: int):
    return await get_planilla(id)

@api_router.post("/planillas/{id}/pagar", response_model=Planilla)
async def pagar_planilla(id: int, cuenta_financiera_id: int = Query(...)):
    """Pay the entire planilla"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            planilla = await conn.fetchrow("SELECT * FROM finanzas2.cont_planilla WHERE id = $1", id)
            if not planilla:
                raise HTTPException(404, "Planilla not found")
            if planilla['estado'] == 'pagada':
                raise HTTPException(400, "Planilla already paid")
            
            # Create pago
            pago_numero = await generate_pago_number(conn, 'egreso')
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas)
                VALUES ($1, 'egreso', $2, $3, $4, $5)
                RETURNING id
            """, pago_numero, datetime.now().date(), cuenta_financiera_id, 
                planilla['total_neto'], f"Pago planilla {planilla['periodo']}")
            
            pago_id = pago['id']
            
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle 
                (pago_id, cuenta_financiera_id, medio_pago, monto)
                VALUES ($1, $2, 'transferencia', $3)
            """, pago_id, cuenta_financiera_id, planilla['total_neto'])
            
            await conn.execute("""
                UPDATE finanzas2.cont_cuenta_financiera 
                SET saldo_actual = saldo_actual - $1 WHERE id = $2
            """, planilla['total_neto'], cuenta_financiera_id)
            
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado)
                VALUES ($1, 'planilla', $2, $3)
            """, pago_id, id, planilla['total_neto'])
            
            # Mark adelantos as descontado
            await conn.execute("""
                UPDATE finanzas2.cont_adelanto_empleado 
                SET descontado = TRUE, planilla_id = $1
                WHERE empleado_id IN (SELECT empleado_id FROM finanzas2.cont_planilla_detalle WHERE planilla_id = $1)
                AND pagado = TRUE AND descontado = FALSE
            """, id)
            
            # Update planilla status
            await conn.execute("""
                UPDATE finanzas2.cont_planilla SET estado = 'pagada', pago_id = $1 WHERE id = $2
            """, pago_id, id)
            
            return await get_planilla(id)

@api_router.delete("/planillas/{id}")
async def delete_planilla(id: int):
    """Delete a planilla (only if in draft status)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if planilla exists and can be deleted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_planilla WHERE id = $1", id
        )
        if not existing:
            raise HTTPException(404, "Planilla no encontrada")
        if existing['estado'] == 'pagada':
            raise HTTPException(400, "No se puede eliminar una planilla pagada")
        
        # Delete detalles first (due to foreign key)
        await conn.execute(
            "DELETE FROM finanzas2.cont_planilla_detalle WHERE planilla_id = $1", id
        )
        # Delete planilla
        await conn.execute(
            "DELETE FROM finanzas2.cont_planilla WHERE id = $1", id
        )
        return {"message": "Planilla eliminada"}

# =====================
# VENTAS POS (Odoo)
# =====================
@api_router.get("/ventas-pos", response_model=List[VentaPOS])
async def list_ventas_pos(
    estado: Optional[str] = None,
    company_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"estado_local = ${idx}")
            params.append(estado)
            idx += 1
        if company_id:
            conditions.append(f"company_id = ${idx}")
            params.append(company_id)
            idx += 1
        if fecha_desde:
            conditions.append(f"date_order >= ${idx}")
            params.append(datetime.combine(fecha_desde, datetime.min.time()))
            idx += 1
        if fecha_hasta:
            conditions.append(f"date_order <= ${idx}")
            params.append(datetime.combine(fecha_hasta, datetime.max.time()))
            idx += 1
        
        query = f"""
            SELECT * FROM finanzas2.cont_venta_pos
            WHERE {' AND '.join(conditions)}
            ORDER BY date_order DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/ventas-pos/sync")
async def sync_ventas_pos(company: str = "ambission", days_back: int = 30):
    """Sync POS orders from Odoo"""
    try:
        odoo = OdooService(company=company)
        if not odoo.authenticate():
            raise HTTPException(401, f"Could not authenticate with Odoo ({company})")
        
        orders = odoo.get_pos_orders(days_back=days_back)
        
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.execute("SET search_path TO finanzas2, public")
            
            synced = 0
            for order in orders:
                # Check if already exists
                existing = await conn.fetchval("""
                    SELECT id FROM finanzas2.cont_venta_pos WHERE odoo_id = $1
                """, order['id'])
                
                if existing:
                    # Update
                    await conn.execute("""
                        UPDATE finanzas2.cont_venta_pos SET
                            date_order = $2,
                            name = $3,
                            partner_id = $4,
                            partner_name = $5,
                            company_id = $6,
                            company_name = $7,
                            amount_total = $8,
                            state = $9,
                            synced_at = NOW()
                        WHERE odoo_id = $1
                    """, order['id'],
                        order.get('date_order'),
                        order.get('name'),
                        order['partner_id'][0] if isinstance(order.get('partner_id'), list) else order.get('partner_id'),
                        order['partner_id'][1] if isinstance(order.get('partner_id'), list) else None,
                        order['company_id'][0] if isinstance(order.get('company_id'), list) else order.get('company_id'),
                        order['company_id'][1] if isinstance(order.get('company_id'), list) else None,
                        order.get('amount_total'),
                        order.get('state'))
                else:
                    # Insert
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_venta_pos 
                        (odoo_id, date_order, name, partner_id, partner_name, company_id, company_name,
                         amount_total, state, synced_at)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, NOW())
                    """, order['id'],
                        order.get('date_order'),
                        order.get('name'),
                        order['partner_id'][0] if isinstance(order.get('partner_id'), list) else order.get('partner_id'),
                        order['partner_id'][1] if isinstance(order.get('partner_id'), list) else None,
                        order['company_id'][0] if isinstance(order.get('company_id'), list) else order.get('company_id'),
                        order['company_id'][1] if isinstance(order.get('company_id'), list) else None,
                        order.get('amount_total'),
                        order.get('state'))
                synced += 1
        
        return {"message": f"Synced {synced} orders from {company}", "synced": synced}
        
    except Exception as e:
        logger.error(f"Error syncing from Odoo: {e}")
        raise HTTPException(500, f"Error syncing: {str(e)}")

@api_router.post("/ventas-pos/{id}/confirmar")
async def confirmar_venta_pos(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        await conn.execute("""
            UPDATE finanzas2.cont_venta_pos SET estado_local = 'confirmada' WHERE id = $1
        """, id)
        return {"message": "Venta confirmada"}

@api_router.post("/ventas-pos/{id}/credito")
async def marcar_credito_venta_pos(id: int, fecha_vencimiento: Optional[date] = None):
    """Mark sale as credit and create CxC"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            venta = await conn.fetchrow("SELECT * FROM finanzas2.cont_venta_pos WHERE id = $1", id)
            if not venta:
                raise HTTPException(404, "Venta not found")
            
            # Create CxC
            cxc = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_cxc 
                (venta_pos_id, monto_original, saldo_pendiente, fecha_vencimiento, estado)
                VALUES ($1, $2, $2, $3, 'pendiente')
                RETURNING id
            """, id, venta['amount_total'], fecha_vencimiento or (datetime.now().date() + timedelta(days=30)))
            
            await conn.execute("""
                UPDATE finanzas2.cont_venta_pos SET estado_local = 'credito', cxc_id = $1, is_credit = TRUE WHERE id = $2
            """, cxc['id'], id)
            
            return {"message": "Venta marcada como crédito", "cxc_id": cxc['id']}

@api_router.post("/ventas-pos/{id}/descartar")
async def descartar_venta_pos(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        await conn.execute("""
            UPDATE finanzas2.cont_venta_pos SET estado_local = 'descartada', is_cancel = TRUE WHERE id = $1
        """, id)
        return {"message": "Venta descartada"}

# =====================
# CXC (Cuentas por Cobrar)
# =====================
@api_router.get("/cxc", response_model=List[CXC])
async def list_cxc(estado: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"cxc.estado = ${idx}")
            params.append(estado)
            idx += 1
        
        query = f"""
            SELECT cxc.*, t.nombre as cliente_nombre
            FROM finanzas2.cont_cxc cxc
            LEFT JOIN finanzas2.cont_tercero t ON cxc.cliente_id = t.id
            WHERE {' AND '.join(conditions)}
            ORDER BY cxc.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

# =====================
# CXP (Cuentas por Pagar)
# =====================
@api_router.get("/cxp")
async def list_cxp(estado: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if estado:
            conditions.append(f"cxp.estado = ${idx}")
            params.append(estado)
            idx += 1
        
        query = f"""
            SELECT cxp.*, t.nombre as proveedor_nombre, fp.numero as factura_numero
            FROM finanzas2.cont_cxp cxp
            LEFT JOIN finanzas2.cont_tercero t ON cxp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON cxp.factura_id = fp.id
            WHERE {' AND '.join(conditions)}
            ORDER BY cxp.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

# =====================
# PRESUPUESTOS
# =====================
@api_router.get("/presupuestos", response_model=List[Presupuesto])
async def list_presupuestos(anio: Optional[int] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if anio:
            rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_presupuesto WHERE anio = $1 ORDER BY version DESC
            """, anio)
        else:
            rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_presupuesto ORDER BY anio DESC, version DESC
            """)
        
        result = []
        for row in rows:
            pres_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT pl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_presupuesto_linea pl
                LEFT JOIN finanzas2.cont_categoria c ON pl.categoria_id = c.id
                WHERE pl.presupuesto_id = $1
            """, row['id'])
            pres_dict['lineas'] = [dict(l) for l in lineas]
            result.append(pres_dict)
        
        return result

@api_router.post("/presupuestos", response_model=Presupuesto)
async def create_presupuesto(data: PresupuestoCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get next version for this year
            version = await conn.fetchval("""
                SELECT COALESCE(MAX(version), 0) + 1 FROM finanzas2.cont_presupuesto WHERE anio = $1
            """, data.anio) or 1
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_presupuesto 
                (nombre, anio, version, estado, notas)
                VALUES ($1, $2, $3, 'borrador', $4)
                RETURNING *
            """, data.nombre, data.anio, version, data.notas)
            
            presupuesto_id = row['id']
            
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_presupuesto_linea 
                    (presupuesto_id, categoria_id, centro_costo_id, linea_negocio_id, mes, monto_presupuestado)
                    VALUES ($1, $2, $3, $4, $5, $6)
                """, presupuesto_id, linea.categoria_id, linea.centro_costo_id,
                    linea.linea_negocio_id, linea.mes, linea.monto_presupuestado)
            
            return await get_presupuesto(presupuesto_id)

async def get_presupuesto(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_presupuesto WHERE id = $1", id)
        if not row:
            raise HTTPException(404, "Presupuesto not found")
        
        pres_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT pl.*, c.nombre as categoria_nombre
            FROM finanzas2.cont_presupuesto_linea pl
            LEFT JOIN finanzas2.cont_categoria c ON pl.categoria_id = c.id
            WHERE pl.presupuesto_id = $1
        """, id)
        pres_dict['lineas'] = [dict(l) for l in lineas]
        
        return pres_dict

@api_router.get("/presupuestos/{id}", response_model=Presupuesto)
async def get_presupuesto_endpoint(id: int):
    return await get_presupuesto(id)

# =====================
# CONCILIACION BANCARIA
# =====================
@api_router.post("/conciliacion/previsualizar-excel")
async def previsualizar_excel_banco(
    file: UploadFile = File(...),
    banco: str = Query(...)
):
    """Preview bank movements from Excel before importing"""
    import io
    from datetime import datetime as dt
    
    try:
        content = await file.read()
        
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Find header row
        header_row = 1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(row):
                row_str = ' '.join([str(c or '') for c in row]).lower()
                if 'fecha' in row_str or 'f. valor' in row_str or 'f. operación' in row_str:
                    header_row = idx
                    break
        
        preview_data = []
        
        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 51, values_only=True):
            if not row or not any(row):
                continue
            
            fecha = None
            descripcion = None
            referencia = None
            monto = None
            
            try:
                if banco == 'BCP':
                    # BCP: Nº, Fecha, Fecha valuta, Descripción operación, Monto, Saldo, Sucursal, Operación-Número
                    fecha = row[1] if len(row) > 1 else None
                    descripcion = row[3] if len(row) > 3 else None
                    monto_val = row[4] if len(row) > 4 else None
                    referencia = row[7] if len(row) > 7 else None
                    
                    if monto_val:
                        monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                            
                elif banco == 'BBVA':
                    # BBVA: N°, F. Operación, F. Valor, Código, Nº. Doc., Concepto, Importe, Oficina
                    # Skip "Saldo Final" rows
                    concepto = row[5] if len(row) > 5 else None
                    if concepto and 'saldo final' in str(concepto).lower():
                        continue
                    
                    fecha = row[1] if len(row) > 1 else None
                    referencia = row[4] if len(row) > 4 else None
                    descripcion = row[5] if len(row) > 5 else None
                    importe = row[6] if len(row) > 6 else None
                    
                    if importe:
                        monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                            
                elif banco == 'IBK':
                    # IBK needs to find the actual data header (Fecha operación, etc.)
                    # Skip metadata rows
                    if not any(row[:2]):
                        continue
                    
                    # Check if this is data row (has date in first or second column)
                    fecha = row[0] if len(row) > 0 else row[1] if len(row) > 1 else None
                    if not fecha:
                        continue
                        
                    descripcion = row[4] if len(row) > 4 else None
                    referencia = row[3] if len(row) > 3 else None
                    cargo = row[6] if len(row) > 6 else None
                    abono = row[7] if len(row) > 7 else None
                    
                    if cargo:
                        cargo_val = float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', ''))
                        monto = -cargo_val
                    elif abono:
                        abono_val = float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', ''))
                        monto = abono_val
                        
                else:
                    # PERSONALIZADO
                    fecha = row[0] if len(row) > 0 else None
                    descripcion = row[1] if len(row) > 1 else None
                    referencia = row[2] if len(row) > 2 else None
                    monto = row[3] if len(row) > 3 else None
                
                # Parse date
                if fecha:
                    if isinstance(fecha, dt):
                        fecha = fecha.date().isoformat()
                    elif hasattr(fecha, 'date'):
                        fecha = fecha.date().isoformat()
                    elif isinstance(fecha, str):
                        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                            try:
                                fecha = dt.strptime(fecha.strip(), fmt).date().isoformat()
                                break
                            except:
                                continue
                
                if not fecha or monto is None:
                    continue
                
                preview_data.append({
                    "fecha": fecha,
                    "banco": banco,  # Use selected bank name
                    "referencia": str(referencia)[:200] if referencia else "",
                    "descripcion": str(descripcion)[:500] if descripcion else "",
                    "monto": float(monto) if monto else 0.0
                })
                
                if len(preview_data) >= 50:
                    break
                    
            except Exception as row_error:
                logger.warning(f"Error parsing row: {row_error}")
                continue
        
        return {
            "preview": preview_data,
            "total_rows": len(preview_data)
        }
        
    except Exception as e:
        logger.error(f"Error previewing Excel: {e}")
        raise HTTPException(500, f"Error al previsualizar: {str(e)}")

@api_router.post("/conciliacion/importar-excel")
async def importar_excel_banco(
    file: UploadFile = File(...),
    cuenta_financiera_id: int = Query(...),
    banco: str = Query(...)
):
    """Import bank movements from Excel - UPSERT based on banco + referencia"""
    import io
    from datetime import datetime as dt
    
    pool = await get_pool()
    
    try:
        content = await file.read()
        
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        async with pool.acquire() as conn:
            await conn.execute("SET search_path TO finanzas2, public")
            
            imported = 0
            updated = 0
            skipped = 0
            
            # Find header row
            header_row = 1
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(row):
                    row_str = ' '.join([str(c or '') for c in row]).lower()
                    if 'fecha' in row_str or 'f. valor' in row_str or 'f. operación' in row_str:
                        header_row = idx
                        break
            
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not any(row):
                    continue
                
                fecha = None
                descripcion = None
                referencia = None
                monto = None
                
                try:
                    if banco == 'BCP':
                        # BCP: Nº, Fecha, Fecha valuta, Descripción operación, Monto, Saldo, Sucursal, Operación-Número
                        fecha = row[1] if len(row) > 1 else None
                        descripcion = row[3] if len(row) > 3 else None
                        monto_val = row[4] if len(row) > 4 else None
                        referencia = row[7] if len(row) > 7 else None
                        
                        if monto_val:
                            monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                                
                    elif banco == 'BBVA':
                        # BBVA: N°, F. Operación, F. Valor, Código, Nº. Doc., Concepto, Importe, Oficina
                        # Skip "Saldo Final" rows
                        concepto = row[5] if len(row) > 5 else None
                        if concepto and 'saldo final' in str(concepto).lower():
                            continue
                        
                        fecha = row[1] if len(row) > 1 else None
                        referencia = row[4] if len(row) > 4 else None
                        descripcion = row[5] if len(row) > 5 else None
                        importe = row[6] if len(row) > 6 else None
                        
                        if importe:
                            monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                                
                    elif banco == 'IBK':
                        # IBK needs to find the actual data rows
                        # Skip metadata rows
                        if not any(row[:2]):
                            continue
                        
                        fecha = row[0] if len(row) > 0 else row[1] if len(row) > 1 else None
                        if not fecha:
                            continue
                            
                        descripcion = row[4] if len(row) > 4 else None
                        referencia = row[3] if len(row) > 3 else None
                        cargo = row[6] if len(row) > 6 else None
                        abono = row[7] if len(row) > 7 else None
                        
                        if cargo:
                            cargo_val = float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', ''))
                            monto = -cargo_val
                        elif abono:
                            abono_val = float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', ''))
                            monto = abono_val
                            
                    else:
                        # PERSONALIZADO
                        fecha = row[0] if len(row) > 0 else None
                        descripcion = row[1] if len(row) > 1 else None
                        referencia = row[2] if len(row) > 2 else None
                        monto = row[3] if len(row) > 3 else None
                    
                    # Parse date
                    if fecha:
                        if isinstance(fecha, dt):
                            fecha = fecha.date()
                        elif hasattr(fecha, 'date'):
                            fecha = fecha.date()
                        elif isinstance(fecha, str):
                            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                                try:
                                    fecha = dt.strptime(fecha.strip(), fmt).date()
                                    break
                                except:
                                    continue
                    
                    # Skip rows without valid date or monto
                    if not fecha or monto is None:
                        continue
                    
                    # Clean fields
                    ref_clean = str(referencia).strip()[:200] if referencia else ''
                    desc_clean = str(descripcion).strip()[:500] if descripcion else ''
                    
                    # Check if exists and if it's already reconciled
                    existing = await conn.fetchrow("""
                        SELECT id, procesado FROM finanzas2.cont_banco_mov_raw 
                        WHERE cuenta_financiera_id = $1 
                          AND banco = $2
                          AND COALESCE(referencia, '') = $3
                          AND fecha = $4
                    """, cuenta_financiera_id, banco, ref_clean, fecha)
                    
                    if existing:
                        if existing['procesado']:
                            skipped += 1
                            continue
                        else:
                            # Update existing record
                            await conn.execute("""
                                UPDATE finanzas2.cont_banco_mov_raw 
                                SET descripcion = $1, monto = $2, banco_excel = $3
                                WHERE id = $4
                            """, desc_clean, monto, banco, existing['id'])
                            updated += 1
                    else:
                        # Insert new record
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_banco_mov_raw 
                            (cuenta_financiera_id, banco, fecha, descripcion, referencia, monto, banco_excel, procesado)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, FALSE)
                        """, cuenta_financiera_id, banco, fecha, desc_clean, 
                            ref_clean if ref_clean else None, monto, banco)
                        imported += 1
                    
                except Exception as row_error:
                    logger.warning(f"Error parsing row: {row_error}")
                    continue
        
        return {
            "message": f"Importados: {imported}, Actualizados: {updated}, Omitidos (ya conciliados): {skipped}",
            "imported": imported,
            "updated": updated,
            "skipped": skipped
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(500, f"Error al importar: {str(e)}")

@api_router.get("/conciliacion/movimientos-banco")
async def list_movimientos_banco(
    cuenta_financiera_id: Optional[int] = None,
    procesado: Optional[bool] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["1=1"]
        params = []
        idx = 1
        
        if cuenta_financiera_id:
            conditions.append(f"cuenta_financiera_id = ${idx}")
            params.append(cuenta_financiera_id)
            idx += 1
        if procesado is not None:
            conditions.append(f"procesado = ${idx}")
            params.append(procesado)
            idx += 1
        
        query = f"""
            SELECT * FROM finanzas2.cont_banco_mov_raw
            WHERE {' AND '.join(conditions)}
            ORDER BY fecha DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.get("/conciliaciones", response_model=List[Conciliacion])
async def list_conciliaciones(cuenta_financiera_id: Optional[int] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if cuenta_financiera_id:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
                WHERE c.cuenta_financiera_id = $1
                ORDER BY c.fecha_fin DESC
            """, cuenta_financiera_id)
        else:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
                ORDER BY c.fecha_fin DESC
            """)
        
        result = []
        for row in rows:
            conc_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT * FROM finanzas2.cont_conciliacion_linea WHERE conciliacion_id = $1
            """, row['id'])
            conc_dict['lineas'] = [dict(l) for l in lineas]
            result.append(conc_dict)
        
        return result

@api_router.post("/conciliacion/conciliar")
async def conciliar_movimientos(
    banco_ids: List[int] = Query(...),
    pago_ids: List[int] = Query(...)
):
    """Mark bank movements and system payments as reconciled"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Mark bank movements as processed
            if banco_ids:
                await conn.execute("""
                    UPDATE finanzas2.cont_banco_mov_raw 
                    SET procesado = TRUE 
                    WHERE id = ANY($1::int[])
                """, banco_ids)
            
            # Mark payments as reconciled (add conciliado field if needed)
            if pago_ids:
                # First ensure the column exists
                try:
                    await conn.execute("""
                        ALTER TABLE finanzas2.cont_pago 
                        ADD COLUMN IF NOT EXISTS conciliado BOOLEAN DEFAULT FALSE
                    """)
                except:
                    pass
                
                await conn.execute("""
                    UPDATE finanzas2.cont_pago 
                    SET conciliado = TRUE 
                    WHERE id = ANY($1::int[])
                """, pago_ids)
        
        return {
            "message": f"Conciliados {len(banco_ids)} movimientos del banco y {len(pago_ids)} del sistema",
            "banco_conciliados": len(banco_ids),
            "sistema_conciliados": len(pago_ids)
        }

@api_router.post("/conciliaciones", response_model=Conciliacion)
async def create_conciliacion(data: ConciliacionCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_conciliacion 
            (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_inicial, saldo_final, notas)
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING *
        """, data.cuenta_financiera_id, data.fecha_inicio, data.fecha_fin,
            data.saldo_inicial, data.saldo_final, data.notas)
        
        result = dict(row)
        result['lineas'] = []
        return result

# =====================
# REPORTES
# =====================
@api_router.get("/reportes/flujo-caja")
async def reporte_flujo_caja(
    fecha_desde: date = Query(...),
    fecha_hasta: date = Query(...)
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT p.fecha, p.tipo, p.monto_total, p.notas,
                   cf.nombre as cuenta
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            WHERE p.fecha BETWEEN $1 AND $2
            ORDER BY p.fecha ASC
        """, fecha_desde, fecha_hasta)
        
        resultado = []
        saldo_acumulado = 0
        
        for row in rows:
            if row['tipo'] == 'ingreso':
                saldo_acumulado += float(row['monto_total'])
            else:
                saldo_acumulado -= float(row['monto_total'])
            
            resultado.append({
                "fecha": row['fecha'].isoformat(),
                "concepto": row['notas'] or row['cuenta'],
                "tipo": row['tipo'],
                "monto": float(row['monto_total']),
                "saldo_acumulado": saldo_acumulado
            })
        
        return resultado

@api_router.get("/reportes/estado-resultados")
async def reporte_estado_resultados(
    fecha_desde: date = Query(...),
    fecha_hasta: date = Query(...)
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Ingresos from ventas POS
        ingresos = await conn.fetchval("""
            SELECT COALESCE(SUM(amount_total), 0) 
            FROM finanzas2.cont_venta_pos 
            WHERE date_order BETWEEN $1 AND $2 AND estado_local != 'descartada'
        """, datetime.combine(fecha_desde, datetime.min.time()), 
            datetime.combine(fecha_hasta, datetime.max.time())) or 0
        
        # Egresos from pagos
        egresos_data = await conn.fetch("""
            SELECT c.nombre as categoria, COALESCE(SUM(gl.importe), 0) as monto
            FROM finanzas2.cont_gasto g
            JOIN finanzas2.cont_gasto_linea gl ON g.id = gl.gasto_id
            LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
            WHERE g.fecha BETWEEN $1 AND $2
            GROUP BY c.nombre
        """, fecha_desde, fecha_hasta)
        
        total_egresos = sum(float(e['monto']) for e in egresos_data)
        
        return {
            "ingresos": [
                {"categoria": "Ventas", "tipo": "ingreso", "monto": float(ingresos)}
            ],
            "egresos": [{"categoria": e['categoria'] or "Sin categoría", "tipo": "egreso", "monto": float(e['monto'])} for e in egresos_data],
            "total_ingresos": float(ingresos),
            "total_egresos": total_egresos,
            "resultado_neto": float(ingresos) - total_egresos
        }

@api_router.get("/reportes/balance-general")
async def reporte_balance_general():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Activos: Bancos y cajas
        activos_bancos = await conn.fetch("""
            SELECT nombre, saldo_actual FROM finanzas2.cont_cuenta_financiera WHERE activo = TRUE
        """)
        
        total_activos = sum(float(a['saldo_actual']) for a in activos_bancos)
        
        # Pasivos: CxP + Letras pendientes
        total_cxp = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) FROM finanzas2.cont_cxp WHERE estado NOT IN ('pagado', 'anulada')
        """) or 0
        
        total_letras = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) FROM finanzas2.cont_letra WHERE estado IN ('pendiente', 'parcial')
        """) or 0
        
        total_pasivos = float(total_cxp) + float(total_letras)
        
        patrimonio = total_activos - total_pasivos
        
        return {
            "activos": [{"cuenta": a['nombre'], "tipo": "activo", "monto": float(a['saldo_actual'])} for a in activos_bancos],
            "pasivos": [
                {"cuenta": "Cuentas por Pagar", "tipo": "pasivo", "monto": float(total_cxp)},
                {"cuenta": "Letras por Pagar", "tipo": "pasivo", "monto": float(total_letras)}
            ],
            "total_activos": total_activos,
            "total_pasivos": total_pasivos,
            "patrimonio": patrimonio
        }

# Include router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
